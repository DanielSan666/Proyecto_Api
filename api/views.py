from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from openpyxl import load_workbook
from django.shortcuts import render, redirect
from .models import  *
from .forms import *
import dateparser
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout
from django.http import HttpResponseForbidden
from django.contrib import messages


def registro(request):
    context = {}
    if request.method == 'POST':
        form = UserRegisterForm(request.POST)
        if form.is_valid():
            form.save()
            username = form.cleaned_data['username']
            messages.success(request, f'Usuario {username} creado')
            return redirect('usuarios')
    else:
        form = UserRegisterForm()
        
    context['form'] = form
    return render(request, 'registro.html', context)

#Usuario
def usuarios(request):
    if request.method == 'GET':
        form = UsuarioForm()
        contexto = {
            'form':form
        }
    else:
        form = UsuarioForm(request.POST)
        contexto = {
            'form':form
        }
        if form.is_valid():
            form.save()
            return redirect('home')
    return render(request, 'usuarios.html', contexto)

def inicio(request):
    return render(request, 'home.html')

#Participantes

def participantes(request):
    partis = Parti.objects.all()
    paginator = Paginator(partis, 10)  # 10 elementos por página
    page = request.GET.get('page',1)

    try:
        partis = paginator.page(page)
    except PageNotAnInteger:
        partis = paginator.page(1)
    except EmptyPage:
        partis = paginator.page(paginator.num_pages)

    contexto = {
        'partis': partis
    }

    return render(request, 'index.html', contexto)

def crearParti(request):
    if request.method == 'GET':
        form = PartiForm()
        contexto = {
            'form':form
        }
    else:
        form = PartiForm(request.POST)
        contexto = {
            'form':form
        }
        if form.is_valid():
            form.save()
            return redirect('index')
    return render(request, 'crear_parti.html', contexto)

def editarParti(request, PartiId):
    parti = Parti.objects.get(PartiId = PartiId)
    if request.method == 'GET':
        form = PartiForm(instance = parti)
        contexto = {
            'form':form
        }
    else:
        form = PartiForm(request.POST, instance = parti)
        contexto = {
            'form':form
        }
        if form.is_valid():
            form.save()
            return redirect('index')
    return render(request, 'editar_parti.html', contexto)

def eliminarParti(request, PartiId):
    parti = Parti.objects.get(PartiId = PartiId)
    parti.delete()
    return redirect('index')


#Api entrenadores
def entrenadores(request):
    entres = Entre.objects.all()
    paginator = Paginator(entres, 10)  # 10 elementos por página
    page = request.GET.get('page',1)

    try:
        entres = paginator.page(page)
    except PageNotAnInteger:
        entres = paginator.page(1)
    except EmptyPage:
        entres = paginator.page(paginator.num_pages)
    contexto = {
        'entres':entres
        }
    return render(request,'entrenadores.html',contexto)

def crearEntre(request):
    if request.method == 'GET':
        form = EnterForm()
    else:
        form = EnterForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('entrenadores')
    
    contexto = {
        'form': form
    }
    return render(request, 'crear_entre.html', contexto)

def editarEntre(request, EntreId):
    entre = Entre.objects.get(EntreId=EntreId)
    if request.method == 'GET':
        form = EnterForm(instance=entre)
        contexto = {
            'form': form
        }
    else:
        form = EnterForm(request.POST, request.FILES, instance=entre)
        contexto = {
            'form': form
        }
        if form.is_valid():
            form.save()
            return redirect('entrenadores')
    return render(request, 'editar_entre.html', contexto)

def eliminarEntre(request, EntreId):
    entre = Entre.objects.get(EntreId = EntreId)
    entre.delete()
    return redirect('entrenadores')



#Api Carga Masiva
def importar_registros(request):
    if request.method == 'POST':
        archivo_excel = request.FILES['archivo_excel']
        wb = load_workbook(archivo_excel)
        hoja = wb.active

        for fila in hoja.iter_rows(min_row=2, values_only=True):
            # Obtener los datos de cada columna en la fila
            matricula = fila[0]
            universidad = fila[1]
            nombre = fila[2]
            grado = fila[3]
            carrera = fila[4]
            curp = fila[5]
            disciplina = fila[6]
            rama = fila[7]
            fecha_ingreso_str = fila[8]  # Obtener el valor de fecha como una cadena desde el archivo Excel

            # Convertir la cadena de fecha a una fecha válida
            fecha_ingreso = dateparser.parse(str(fecha_ingreso_str))


            ciclo_escolar = fila[9]
            sexo = fila[10]
            nss = fila[11]

            # Crear un nuevo objeto Parti y guardar los datos
            parti = Parti(
                matricula=matricula,
                universidad=universidad,
                nombre=nombre,
                grado=grado,
                carrera=carrera,
                curp=curp,
                disciplina=disciplina,
                rama=rama,
                FechaIngre=fecha_ingreso,  # Actualizar el nombre del campo
                cicloescolar=ciclo_escolar,  # Actualizar el nombre del campo
                sexo=sexo,
                nss=nss
            )
            parti.save()

        return redirect('index')  # Redirigir a la página de entrenadores después de importar los registros

    else:
        form = ImportForm()

    return render(request, 'importar.html', {'form': form})


#Universidades
def universidades(request):
    unis = Uni.objects.all()
    paginator = Paginator(unis, 10)  # 10 elementos por página
    page = request.GET.get('page',1)

    try:
        unis = paginator.page(page)
    except PageNotAnInteger:
        unis = paginator.page(1)
    except EmptyPage:
        unis = paginator.page(paginator.num_pages)

    contexto = {
        'unis': unis
    }

    return render(request, 'universidades.html', contexto)

def crearUni(request):
    if request.method == 'GET':
        form = UniForm()
    else:
        form = UniForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('universidades')
    
    contexto = {
        'form': form
    }
    return render(request, 'crear_uni.html', contexto)

    
def editarUni(request, uniId):
    uni = Uni.objects.get(uniId=uniId)
    if request.method == 'GET':
        form = UniForm(instance=uni)
        contexto = {
            'form': form
        }
    else:
        form = UniForm(request.POST, request.FILES, instance=uni)
        contexto = {
            'form': form
        }
        if form.is_valid():
            form.save()
            return redirect('universidades')
    return render(request, 'editar_uni.html', contexto)

def eliminarUni(request, uniId):
    uni = Uni.objects.get(uniId = uniId)
    uni.delete()
    return redirect('universidades')


def salir(request):
    logout(request)
    return redirect('login')


#Coordinadores
def coordinadores(request):
    coordis = Coordi.objects.all()
    paginator = Paginator(coordis, 10)  # 10 elementos por página
    page = request.GET.get('page',1)

    try:
        coordis = paginator.page(page)
    except PageNotAnInteger:
        coordis = paginator.page(1)
    except EmptyPage:
        coordis = paginator.page(paginator.num_pages)

    contexto = {
        'coordis': coordis
    }


    return render(request, 'coordinadores.html', contexto)


def crearCoordi(request):
    if request.method == 'GET':
        form = CoordiForm()
        contexto = {
            'form': form
        }
    else:
        form = CoordiForm(request.POST, request.FILES)
        contexto = {
            'form': form
        }
        if form.is_valid():
            coordi = form.save(commit=False)
            coordi.save()
            return redirect('coordinadores')
    return render(request, 'crear_coordi.html', contexto)



def editarCoordi(request, corId):
    coordi = Coordi.objects.get(corId=corId)
    if request.method == 'GET':
        form = UniForm(instance=coordi)
        contexto = {
            'form': form
        }
    else:
        form = CoordiForm(request.POST, request.FILES, instance=coordi)
        contexto = {
            'form': form
        }
        if form.is_valid():
            form.save()
            return redirect('coordinadores')
    return render(request, 'editar_coordi.html', contexto)


def eliminarCoordi(request, corId):
    coordi = Coordi.objects.get(corId = corId)
    coordi.delete()
    return redirect('coordinadores')


#Api Asistentes
def asistente(request):
    asist = Asis.objects.all()
    paginator = Paginator(asist, 10)  # 10 elementos por página
    page = request.GET.get('page',1)

    try:
        asist = paginator.page(page)
    except PageNotAnInteger:
        asist = paginator.page(1)
    except EmptyPage:
        asist = paginator.page(paginator.num_pages)

    contexto = {
        'asist': asist
    }


    return render(request, 'asistentes.html', contexto)

def crearAsis(request):
    if request.method == 'GET':
        form = AsisForm()
        contexto = {
            'form': form
        }
    else:
        form = AsisForm(request.POST, request.FILES)
        contexto = {
            'form': form
        }
        if form.is_valid():
            coordi = form.save(commit=False)
            coordi.save()
            return redirect('asistentes')
    return render(request, 'crear_asis.html', contexto)



def editarAsis(request, AsId):
    asis = Asis.objects.get(AsId = AsId)
    if request.method == 'GET':
        form = AsisForm(instance=asis)
        contexto = {
            'form': form
        }
    else:
        form = AsisForm(request.POST, request.FILES, instance=asis)
        contexto = {
            'form': form
        }
        if form.is_valid():
            form.save()
            return redirect('asistentes')
    return render(request, 'editar_asis.html', contexto)


def eliminarAsis(request, AsId):
    asis = Asis.objects.get(AsId = AsId)
    asis.delete()
    return redirect('asistentes')

#Api Medicos
def medicos(request):
    medic = Medi.objects.all()
    paginator = Paginator(medic, 10)  # 10 elementos por página
    page = request.GET.get('page',1)

    try:
        medic = paginator.page(page)
    except PageNotAnInteger:
        medic = paginator.page(1)
    except EmptyPage:
        medic = paginator.page(paginator.num_pages)

    contexto = {
        'medic': medic
    }


    return render(request, 'medicos.html', contexto)

def crearMedic(request):
    if request.method == 'GET':
        form = MedicForm()
    else:
        form = MedicForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('medicos')
    
    contexto = {
        'form': form
    }
    return render(request, 'crear_medic.html', contexto)

    
def editarMedic(request, MedId):
    medi = Medi.objects.get(MedId=MedId)
    if request.method == 'GET':
        form = MedicForm(instance=medi)
        contexto = {
            'form': form
        }
    else:
        form = MedicForm(request.POST, request.FILES, instance=medi)
        contexto = {
            'form': form
        }
        if form.is_valid():
            form.save()
            return redirect('medicos')
    return render(request, 'editar_medic.html', contexto)

def eliminarMedic(request, MedId):
    medi = Medi.objects.get(MedId = MedId)
    medi.delete()
    return redirect('medicos')


#Api Staff
def staff(request):
    staf = Staff.objects.all()
    paginator = Paginator(staf, 10)  # 10 elementos por página
    page = request.GET.get('page',1)

    try:
        staf = paginator.page(page)
    except PageNotAnInteger:
        staf = paginator.page(1)
    except EmptyPage:
        staf = paginator.page(paginator.num_pages)

    contexto = {
        'staf': staf
    }

    return render(request, 'staff.html', contexto)

def crearStaff(request):
    if request.method == 'GET':
        form = StaffForm()
    else:
        form = StaffForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('staff')
    
    contexto = {
        'form': form
    }
    return render(request, 'crear_staff.html', contexto)

    
def editarStaff(request, StId):
    sta = Staff.objects.get(StId=StId)
    if request.method == 'GET':
        form = StaffForm(instance=sta)
        contexto = {
            'form': form
        }
    else:
        form = StaffForm(request.POST, request.FILES, instance=sta)
        contexto = {
            'form': form
        }
        if form.is_valid():
            form.save()
            return redirect('staff')
    return render(request, 'editar_staff.html', contexto)

def eliminarStaff(request, StId):
    sta = Staff.objects.get(StId = StId)
    sta.delete()
    return redirect('staff')